"""
对四个USD投入产出表进行 min-max 归一化: x' = (x - min) / (max - min)

使用全局 min/max（跨4个表所有数值单元格），保留国家间、年份间的相对大小关系。
"""

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from copy import copy
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FILES = [
    "CHN2017_5sector_USD.xlsx",
    "USA2017_5sector_USD.xlsx",
    "CHN2018_5sector_USD.xlsx",
    "USA2018_5sector_USD.xlsx",
]

HEADER_ROWS = 5
LABEL_COLS = 4


def copy_cell_style(src, dst):
    if src.font:
        dst.font = copy(src.font)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.border:
        dst.border = copy(src.border)
    if src.fill:
        dst.fill = copy(src.fill)


def collect_global_min_max():
    """Scan all 4 tables to find global min and max of numeric data cells."""
    global_min = float('inf')
    global_max = float('-inf')

    for f in FILES:
        wb = openpyxl.load_workbook(os.path.join(BASE_DIR, f), data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row_idx in range(HEADER_ROWS + 1, ws.max_row + 1):
            for col_idx in range(LABEL_COLS + 1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue
                v = cell.value
                if isinstance(v, (int, float)):
                    global_min = min(global_min, v)
                    global_max = max(global_max, v)

    return global_min, global_max


def normalize_file(input_file, output_file, g_min, g_max):
    """Normalize one IO table and save."""
    in_path = os.path.join(BASE_DIR, input_file)
    out_path = os.path.join(BASE_DIR, output_file)
    rng = g_max - g_min

    wb_src = openpyxl.load_workbook(in_path, data_only=True)
    wb_dst = openpyxl.Workbook()
    wb_dst.remove(wb_dst.active)

    for sheet_name in wb_src.sheetnames:
        ws_src = wb_src[sheet_name]
        ws_dst = wb_dst.create_sheet(title=sheet_name)

        # Copy merged cells
        for merge_range in ws_src.merged_cells.ranges:
            ws_dst.merge_cells(str(merge_range))

        # Copy column widths / row heights
        for col_idx in range(1, ws_src.max_column + 1):
            cl = get_column_letter(col_idx)
            if ws_src.column_dimensions[cl].width:
                ws_dst.column_dimensions[cl].width = ws_src.column_dimensions[cl].width
        for row_idx in range(1, ws_src.max_row + 1):
            if ws_src.row_dimensions[row_idx].height:
                ws_dst.row_dimensions[row_idx].height = ws_src.row_dimensions[row_idx].height

        is_data_sheet = ("IO" in sheet_name or "非竞争" in sheet_name)

        for row_idx in range(1, ws_src.max_row + 1):
            for col_idx in range(1, ws_src.max_column + 1):
                src_cell = ws_src.cell(row=row_idx, column=col_idx)
                dst_cell = ws_dst.cell(row=row_idx, column=col_idx)

                if isinstance(src_cell, MergedCell) or isinstance(dst_cell, MergedCell):
                    continue

                v = src_cell.value

                if (is_data_sheet
                        and row_idx > HEADER_ROWS
                        and col_idx > LABEL_COLS
                        and isinstance(v, (int, float))):
                    dst_cell.value = (v - g_min) / rng
                    dst_cell.number_format = '0.000000'
                else:
                    dst_cell.value = v

                copy_cell_style(src_cell, dst_cell)

        # Update title / unit rows for data sheet
        if is_data_sheet:
            title_cell = ws_dst.cell(row=1, column=1)
            old_title = title_cell.value or ""
            if "归一化" not in old_title:
                title_cell.value = old_title + "【min-max 归一化】"

            unit_cell = ws_dst.cell(row=2, column=1)
            unit_cell.value = (
                f"归一化: (x - min) / (max - min)    "
                f"[全局 min={g_min:,.3f}, max={g_max:,.3f} 百万美元]"
            )

    wb_dst.save(out_path)


def verify(output_file, g_min, g_max):
    """Print TI row for verification."""
    rng = g_max - g_min
    wb = openpyxl.load_workbook(os.path.join(BASE_DIR, output_file), data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        if row[3].value == "TI":
            vals = [row[c].value for c in range(4, 9)]
            labels = ["A01", "RUP", "RMID", "ELEC/CIT", "TEQ"]
            print(f"  TI: " + "  ".join(
                f"{l}={v:.6f}" for l, v in zip(labels, vals) if v is not None
            ))
            break
    # Also check a known value
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        if row[3].value == "TVA":
            vals = [row[c].value for c in range(4, 9)]
            print(f"  TVA: " + "  ".join(
                f"{v:.6f}" for v in vals if v is not None
            ))
            break


def main():
    print("=" * 70)
    print("Min-Max 归一化: x' = (x - min) / (max - min)")
    print("=" * 70)

    g_min, g_max = collect_global_min_max()
    print(f"\n全局统计:")
    print(f"  min = {g_min:>14,.3f} 百万美元")
    print(f"  max = {g_max:>14,.3f} 百万美元")
    print(f"  range = {g_max - g_min:>14,.3f}")

    for f in FILES:
        out = f.replace("_USD.xlsx", "_normalized.xlsx")
        print(f"\n--- {f} → {out} ---")
        normalize_file(f, out, g_min, g_max)
        verify(out, g_min, g_max)

    print(f"\n{'='*70}")
    print("输出文件:")
    for f in FILES:
        out = f.replace("_USD.xlsx", "_normalized.xlsx")
        print(f"  io_excel/{out}")

    # Summary comparison
    print(f"\n{'='*70}")
    print("总投入归一化值对比:")
    print(f"{'文件':<28} {'A01':>10} {'RUP':>10} {'RMID':>10} {'ELEC':>10} {'TEQ':>10}")
    print("-" * 80)
    for f in FILES:
        out = f.replace("_USD.xlsx", "_normalized.xlsx")
        wb = openpyxl.load_workbook(os.path.join(BASE_DIR, out), data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            if row[3].value == "TI":
                vals = [row[c].value for c in range(4, 9)]
                label = out.replace("_5sector_normalized.xlsx", "")
                print(f"{label:<28} " + " ".join(
                    f"{v:>10.6f}" for v in vals if v is not None
                ))
                break


if __name__ == "__main__":
    main()
