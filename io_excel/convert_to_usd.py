"""
将四个五部门投入产出表统一为美元计价（百万美元），并添加资本要素价格行。

汇率：
  2017年：1 USD = 6.7518 CNY
  2018年：1 USD = 6.6174 CNY

资本要素价格：
  CN 2017: 3.5798%    CN 2018: 3.6222%
  US 2017: 2.33%      US 2018: 2.91%
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
import os

# --- 参数 ---
EXCHANGE_RATES = {2017: 6.7518, 2018: 6.6174}
CAPITAL_PRICES = {
    ("CN", 2017): 0.035798,
    ("CN", 2018): 0.036222,
    ("US", 2017): 0.0233,
    ("US", 2018): 0.0291,
}

# 文件映射: (输入文件, 国家, 年份, 输出文件)
FILES = [
    ("CHN2017_5sector_CNlike.xlsx", "CN", 2017, "CHN2017_5sector_USD.xlsx"),
    ("USA2017_5sector_CNlike.xlsx", "US", 2017, "USA2017_5sector_USD.xlsx"),
    ("2018_CN_五部门_仿42版式_路径A.xlsx", "CN", 2018, "CHN2018_5sector_USD.xlsx"),
    ("2018_US_五部门_仿中表_路径A.xlsx", "US", 2018, "USA2018_5sector_USD.xlsx"),
]

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def copy_cell_style(src, dst):
    """Copy style from src cell to dst cell."""
    if src.font:
        dst.font = copy(src.font)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.border:
        dst.border = copy(src.border)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.number_format:
        dst.number_format = src.number_format


def convert_sheet(ws_src, ws_dst, country, year):
    """
    Convert IO table sheet to USD (百万美元).

    Chinese tables: 万元 → 百万美元 = value / (exchange_rate * 100)
    US tables: already 百万美元, no conversion needed.

    Strategy: copy all cells, convert numeric values in the data area.
    """
    from openpyxl.cell.cell import MergedCell

    is_china = (country == "CN")
    rate = EXCHANGE_RATES[year]
    conversion_factor = 1.0 / (rate * 100.0) if is_china else 1.0
    cap_price = CAPITAL_PRICES[(country, year)]

    max_row = ws_src.max_row
    max_col = ws_src.max_column

    # Copy merged cells first
    for merge_range in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merge_range))

    # Copy column widths
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        if ws_src.column_dimensions[col_letter].width:
            ws_dst.column_dimensions[col_letter].width = ws_src.column_dimensions[col_letter].width

    # Copy row heights
    for row_idx in range(1, max_row + 1):
        if ws_src.row_dimensions[row_idx].height:
            ws_dst.row_dimensions[row_idx].height = ws_src.row_dimensions[row_idx].height

    HEADER_ROWS = 5
    LABEL_COLS = 4

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            src_cell = ws_src.cell(row=row_idx, column=col_idx)
            dst_cell = ws_dst.cell(row=row_idx, column=col_idx)

            # Skip MergedCell (read-only slave cells)
            if isinstance(src_cell, MergedCell) or isinstance(dst_cell, MergedCell):
                continue

            val = src_cell.value

            if (row_idx > HEADER_ROWS and col_idx > LABEL_COLS
                    and isinstance(val, (int, float)) and val is not None):
                dst_cell.value = val * conversion_factor
                dst_cell.number_format = '#,##0.000'
            else:
                dst_cell.value = val

            copy_cell_style(src_cell, dst_cell)

    # Update title row (row 1) - add USD note
    title_cell = ws_dst.cell(row=1, column=1)
    old_title = title_cell.value or ""
    if is_china:
        title_cell.value = old_title + "【已折算为百万美元】"

    # Update unit row (row 2)
    unit_cell = ws_dst.cell(row=2, column=1)
    unit_cell.value = f"单位：百万美元 (Millions of USD)    [汇率: 1 USD = {rate} CNY]" if is_china else "单位：百万美元 (Millions of USD)"

    # --- Add capital factor price row ---
    # Find the last row (总投入/TI row)
    ti_row = None
    for row_idx in range(1, max_row + 1):
        cell_val = ws_src.cell(row=row_idx, column=4).value
        if cell_val == "TI":
            ti_row = row_idx
            break

    if ti_row:
        # Insert capital factor price info after 增加值合计 (TVA) row
        # Find TVA row
        tva_row = None
        for row_idx in range(1, max_row + 1):
            cell_val = ws_src.cell(row=row_idx, column=4).value
            if cell_val == "TVA":
                tva_row = row_idx
                break

        # Add capital price row after the last row
        new_row = max_row + 2
        ws_dst.cell(row=new_row, column=1).value = "资本要素价格"
        ws_dst.cell(row=new_row, column=1).font = Font(bold=True)
        ws_dst.cell(row=new_row, column=3).value = f"{cap_price*100:.4f}%"
        ws_dst.cell(row=new_row, column=3).font = Font(bold=True, color="FF0000")

        note = "即资本回报率/利率，用于模型中资本要素的定价"
        ws_dst.cell(row=new_row, column=5).value = note

    return conversion_factor


def convert_annotation_sheet(ws_src, ws_dst):
    """Copy annotation sheet as-is."""
    from openpyxl.cell.cell import MergedCell

    for merge_range in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merge_range))

    for row in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row,
                                 max_col=ws_src.max_column):
        for cell in row:
            dst_cell = ws_dst.cell(row=cell.row, column=cell.column)
            if isinstance(cell, MergedCell) or isinstance(dst_cell, MergedCell):
                continue
            dst_cell.value = cell.value
            copy_cell_style(cell, dst_cell)


def process_file(input_file, country, year, output_file):
    """Process one IO table file."""
    in_path = os.path.join(BASE_DIR, input_file)
    out_path = os.path.join(BASE_DIR, output_file)

    print(f"\n{'='*60}")
    print(f"Processing: {input_file}")
    print(f"  Country: {country}, Year: {year}")

    wb_src = openpyxl.load_workbook(in_path, data_only=True)
    wb_dst = openpyxl.Workbook()

    # Remove default sheet
    wb_dst.remove(wb_dst.active)

    for sheet_name in wb_src.sheetnames:
        ws_src = wb_src[sheet_name]
        ws_dst = wb_dst.create_sheet(title=sheet_name)

        if "IO" in sheet_name or "投入产出" in sheet_name or "非竞争" in sheet_name:
            factor = convert_sheet(ws_src, ws_dst, country, year)
            is_china = (country == "CN")
            if is_china:
                print(f"  Sheet '{sheet_name}': converted 万元 → 百万美元 (×{factor:.10f})")
            else:
                print(f"  Sheet '{sheet_name}': already in 百万美元, no conversion")
            print(f"  Capital factor price: {CAPITAL_PRICES[(country, year)]*100:.4f}%")
        else:
            convert_annotation_sheet(ws_src, ws_dst)
            print(f"  Sheet '{sheet_name}': copied (annotation)")

    wb_dst.save(out_path)
    print(f"  Saved: {output_file}")

    # --- Verification: print summary ---
    print(f"\n  --- Verification (总投入 TI row) ---")
    wb_check = openpyxl.load_workbook(out_path, data_only=True)
    ws_check = wb_check[wb_check.sheetnames[0]]
    for row in ws_check.iter_rows(min_row=1, max_row=ws_check.max_row, values_only=False):
        if row[3].value == "TI":
            sectors = []
            for col_idx in range(4, 9):  # columns E-I (sectors 1-5)
                v = row[col_idx].value
                if v is not None:
                    sectors.append(f"{v:,.3f}")
            print(f"  TI (百万美元): {', '.join(sectors)}")
            break


def main():
    print("=" * 60)
    print("投入产出表货币统一工具")
    print("目标：全部折算为 百万美元 (Millions of USD)")
    print("=" * 60)

    for input_file, country, year, output_file in FILES:
        process_file(input_file, country, year, output_file)

    print(f"\n{'='*60}")
    print("全部完成！生成文件：")
    for _, _, _, output_file in FILES:
        print(f"  io_excel/{output_file}")

    # Cross-check: print all TI for comparison
    print(f"\n{'='*60}")
    print("总投入对比 (百万美元):")
    print(f"{'文件':<30} {'A01':>12} {'RUP':>12} {'RMID':>12} {'ELEC/CIT':>12} {'TEQ':>12} {'合计':>14}")
    print("-" * 100)

    for _, country, year, output_file in FILES:
        out_path = os.path.join(BASE_DIR, output_file)
        wb = openpyxl.load_workbook(out_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            if row[3].value == "TI":
                vals = []
                total = 0
                for col_idx in range(4, 9):
                    v = row[col_idx].value
                    if v is not None:
                        vals.append(v)
                        total += v
                    else:
                        vals.append(0)
                label = f"{country}{year}"
                print(f"{label:<30} {vals[0]:>12,.1f} {vals[1]:>12,.1f} {vals[2]:>12,.1f} {vals[3]:>12,.1f} {vals[4]:>12,.1f} {total:>14,.1f}")
                break


if __name__ == "__main__":
    main()
