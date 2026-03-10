"""
从投入产出表计算模型参数（A/B/C三类）。

策略:
  - 比值类参数（alpha_ij, theta_ij, beta_j, theta_cj, gamma_ij, gamma_cj）
    从归一化前的 USD 数据计算，保持经济学含义
  - 绝对量参数（ExportValue_j, factor_params, E_j）
    从归一化后数据计算

资本要素价格:
  CN 2017: 3.5798%    CN 2018: 3.6222%
  US 2017: 2.33%      US 2018: 2.91%
"""

from __future__ import annotations

import json
import os
from dataclasses import dataclass
from typing import Dict, List

import numpy as np
import openpyxl
from openpyxl.cell.cell import MergedCell

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR = os.path.join(os.path.dirname(BASE_DIR), "io_excel")

# ──────────────────────────────────────────────
# 配置
# ──────────────────────────────────────────────

CAPITAL_PRICES = {
    ("CN", 2017): 0.035798,
    ("CN", 2018): 0.036222,
    ("US", 2017): 0.0233,
    ("US", 2018): 0.0291,
}

# (归一化文件, USD文件, 国家, 年份)
FILES = [
    ("CHN2017_5sector_normalized.xlsx", "CHN2017_5sector_USD.xlsx", "CN", 2017),
    ("USA2017_5sector_normalized.xlsx", "USA2017_5sector_USD.xlsx", "US", 2017),
    ("CHN2018_5sector_normalized.xlsx", "CHN2018_5sector_USD.xlsx", "CN", 2018),
    ("USA2018_5sector_normalized.xlsx", "USA2018_5sector_USD.xlsx", "US", 2018),
]

ELEC_VARIANTS = {"ELEC", "CIT"}
VA_CODES = {"VA001": "劳动者报酬", "VA002": "生产税净额",
            "VA003": "固定资产折旧", "VA004": "营业盈余"}

N_PERIODS = 10
RHO_CJ_DEFAULT = 0.95
RHO_IJ_DEFAULT = 0.95   # 生产侧 CES 形状参数（与消费侧一致）
IMPORT_COST_DEFAULT = 1.0  # 冰山贸易成本（1.0 = 无额外成本）


# ──────────────────────────────────────────────
# IO 表解析器
# ──────────────────────────────────────────────

@dataclass
class ParsedIOTable:
    """从 Excel 解析后的 IO 表结构化数据。"""
    country: str
    year: int
    sectors: List[str]
    raw_sectors: List[str]
    Z_dom: np.ndarray       # [5×5], Z_dom[j,i] = supplier j → user i
    Z_imp: np.ndarray       # [5×5]
    FD_dom: Dict[str, np.ndarray]  # {TC, FU201, FU202, EX} → [5]
    FD_imp: Dict[str, np.ndarray]
    VA: Dict[str, np.ndarray]      # {VA001..VA004} → [5]
    TI: np.ndarray          # [5]
    GO_dom: np.ndarray      # [5]
    GO_imp: np.ndarray      # [5]


def _safe_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _cell_val(ws, row, col) -> float:
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return 0.0
    return _safe_float(cell.value)


def _normalize_sector_code(code: str) -> str:
    if code in ELEC_VARIANTS:
        return "ELEC_CIT"
    return code


def parse_excel_io_table(filepath: str, country: str, year: int) -> ParsedIOTable:
    """解析 Excel IO 表（支持 USD 和归一化两种格式）。"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    max_row = ws.max_row
    max_col = ws.max_column

    row_info = {}

    for r in range(1, max_row + 1):
        code_cell = ws.cell(row=r, column=4)
        if isinstance(code_cell, MergedCell):
            continue
        code = code_cell.value
        if code is None:
            continue
        code = str(code).strip()

        if code in ("A01", "RUP", "RMID", "ELEC", "CIT", "TEQ"):
            col2_val = ""
            c2 = ws.cell(row=r, column=2)
            if not isinstance(c2, MergedCell) and c2.value is not None:
                col2_val = str(c2.value).strip()
            if not col2_val:
                for scan_r in range(r - 1, 0, -1):
                    c2s = ws.cell(row=scan_r, column=2)
                    if isinstance(c2s, MergedCell):
                        continue
                    if c2s.value is not None:
                        col2_val = str(c2s.value).strip()
                        break

            is_import = "进口" in col2_val or "进" in col2_val
            normalized = _normalize_sector_code(code)

            if is_import:
                row_info[r] = ("import_sector", normalized, code)
            else:
                row_info[r] = ("domestic_sector", normalized, code)

        elif code in VA_CODES:
            row_info[r] = ("va", code, code)
        elif code == "TVA":
            row_info[r] = ("tva", code, code)
        elif code == "TI":
            row_info[r] = ("ti", code, code)
        elif code == "TII":
            row_info[r] = ("tii", code, code)

    raw_sector_order = []
    sector_order = []
    for r in sorted(row_info.keys()):
        rtype, code, orig = row_info[r]
        if rtype == "domestic_sector" and code not in sector_order:
            sector_order.append(code)
            raw_sector_order.append(orig)

    n = len(sector_order)
    assert n == 5, f"Expected 5 sectors, found {n}: {sector_order}"

    col_map = {}
    fd_col_map = {}
    for c in range(1, max_col + 1):
        cell = ws.cell(row=5, column=c)
        if isinstance(cell, MergedCell):
            continue
        val = cell.value
        if val is None:
            continue
        val = str(val).strip()
        norm_val = _normalize_sector_code(val)
        if norm_val in set(sector_order):
            col_map[norm_val] = c
        elif val == "TC":
            fd_col_map["TC"] = c
        elif val == "FU201":
            fd_col_map["FU201"] = c
        elif val == "FU202":
            fd_col_map["FU202"] = c
        elif val == "EX":
            fd_col_map["EX"] = c
        elif val == "GO/IM":
            fd_col_map["GO_IM"] = c
        elif val == "TIU":
            fd_col_map["TIU"] = c

    Z_dom = np.zeros((n, n))
    Z_imp = np.zeros((n, n))
    FD_dom = {k: np.zeros(n) for k in ["TC", "FU201", "FU202", "EX"]}
    FD_imp = {k: np.zeros(n) for k in ["TC", "FU201", "FU202", "EX"]}
    VA = {k: np.zeros(n) for k in VA_CODES}
    TI = np.zeros(n)
    GO_dom = np.zeros(n)
    GO_imp = np.zeros(n)

    for r, (rtype, code, orig) in sorted(row_info.items()):
        if rtype == "domestic_sector":
            j = sector_order.index(code)
            for i_idx, s_i in enumerate(sector_order):
                if s_i in col_map:
                    Z_dom[j, i_idx] = _cell_val(ws, r, col_map[s_i])
            for fd_name in ["TC", "FU201", "FU202", "EX"]:
                if fd_name in fd_col_map:
                    FD_dom[fd_name][j] = _cell_val(ws, r, fd_col_map[fd_name])
            if "GO_IM" in fd_col_map:
                GO_dom[j] = _cell_val(ws, r, fd_col_map["GO_IM"])

        elif rtype == "import_sector":
            j = sector_order.index(code)
            for i_idx, s_i in enumerate(sector_order):
                if s_i in col_map:
                    Z_imp[j, i_idx] = _cell_val(ws, r, col_map[s_i])
            for fd_name in ["TC", "FU201", "FU202", "EX"]:
                if fd_name in fd_col_map:
                    FD_imp[fd_name][j] = _cell_val(ws, r, fd_col_map[fd_name])
            if "GO_IM" in fd_col_map:
                GO_imp[j] = _cell_val(ws, r, fd_col_map["GO_IM"])

        elif rtype == "va":
            for i_idx, s_i in enumerate(sector_order):
                if s_i in col_map:
                    VA[code][i_idx] = _cell_val(ws, r, col_map[s_i])

        elif rtype == "ti":
            for i_idx, s_i in enumerate(sector_order):
                if s_i in col_map:
                    TI[i_idx] = _cell_val(ws, r, col_map[s_i])

    return ParsedIOTable(
        country=country, year=year,
        sectors=sector_order, raw_sectors=raw_sector_order,
        Z_dom=Z_dom, Z_imp=Z_imp,
        FD_dom=FD_dom, FD_imp=FD_imp,
        VA=VA, TI=TI,
        GO_dom=GO_dom, GO_imp=GO_imp,
    )


# ──────────────────────────────────────────────
# A 类参数计算（从 USD 原始数据，保持经济学比值含义）
# ──────────────────────────────────────────────

def compute_alpha_ij(t: ParsedIOTable) -> np.ndarray:
    """alpha[i,j] = Z_total[j,i] / TI[i]"""
    Z_total = t.Z_dom + t.Z_imp
    n = len(t.sectors)
    alpha = np.zeros((n, n))
    for i in range(n):
        if t.TI[i] > 0:
            for j in range(n):
                alpha[i, j] = Z_total[j, i] / t.TI[i]
    return alpha


def compute_theta_ij(t: ParsedIOTable) -> np.ndarray:
    """theta[i,j] = Z_dom[j,i] / Z_total[j,i], fill NaN→1.0"""
    Z_total = t.Z_dom + t.Z_imp
    n = len(t.sectors)
    theta = np.ones((n, n))
    for i in range(n):
        for j in range(n):
            total = Z_total[j, i]
            if abs(total) > 1e-15:
                theta[i, j] = t.Z_dom[j, i] / total
    return np.clip(theta, 0.0, 1.0)


def compute_beta_j(t: ParsedIOTable) -> np.ndarray:
    """beta[j] = TC_total[j] / sum(TC_total)"""
    TC_total = t.FD_dom["TC"] + t.FD_imp["TC"]
    TC_positive = np.maximum(TC_total, 0.0)
    total = TC_positive.sum()
    if total > 0:
        return TC_positive / total
    return np.ones(len(t.sectors)) / len(t.sectors)


def compute_theta_cj(t: ParsedIOTable) -> np.ndarray:
    """theta_c[j] = TC_dom[j] / TC_total[j], fill NaN→1.0"""
    TC_total = t.FD_dom["TC"] + t.FD_imp["TC"]
    n = len(t.sectors)
    theta_c = np.ones(n)
    for j in range(n):
        if abs(TC_total[j]) > 1e-15:
            theta_c[j] = t.FD_dom["TC"][j] / TC_total[j]
    return np.clip(theta_c, 0.0, 1.0)


# ──────────────────────────────────────────────
# B 类参数计算（绝对量从归一化数据）
# ──────────────────────────────────────────────

def compute_factor_params(t: ParsedIOTable, cap_price: float) -> Dict[str, float]:
    labor_income = t.VA["VA001"].sum()
    capital_income = t.VA["VA003"].sum() + t.VA["VA004"].sum()
    w = 1.0
    r = cap_price
    L = labor_income / w
    K = capital_income / r if r > 0 else 0.0
    return {"w": w, "r": r, "K": K, "L": L}


def compute_export_value_j(t: ParsedIOTable) -> np.ndarray:
    return np.maximum(t.FD_dom["EX"], 0.0)


# ──────────────────────────────────────────────
# C 类参数
# ──────────────────────────────────────────────

def compute_gamma_ij(theta_ij: np.ndarray) -> np.ndarray:
    return np.clip(theta_ij, 1e-6, 1 - 1e-6)


def compute_gamma_cj(theta_cj: np.ndarray) -> np.ndarray:
    return np.clip(theta_cj, 1e-6, 1 - 1e-6)


# ──────────────────────────────────────────────
# CSV 输出
# ──────────────────────────────────────────────

def save_matrix_csv(filepath: str, matrix: np.ndarray, sectors: List[str],
                    row_label: str = "user_sector_i", col_label: str = "input_sector_j"):
    with open(filepath, "w") as f:
        f.write(f"{row_label},{','.join(sectors)}\n")
        for i, s in enumerate(sectors):
            vals = ",".join(str(matrix[i, j]) for j in range(len(sectors)))
            f.write(f"{s},{vals}\n")


def save_vector_csv(filepath: str, vector: np.ndarray, sectors: List[str],
                    index_name: str = "sector_j", col_name: str = "value"):
    with open(filepath, "w") as f:
        f.write(f"{index_name},{col_name}\n")
        for i, s in enumerate(sectors):
            f.write(f"{s},{vector[i]}\n")


def save_time_series_csv(filepath: str, n_periods: int, sectors: List[str],
                         value: float = 1.0):
    with open(filepath, "w") as f:
        f.write(f"period_t,{','.join(sectors)}\n")
        row = ",".join(str(value) for _ in sectors)
        for t in range(n_periods):
            f.write(f"{t},{row}\n")


# ──────────────────────────────────────────────
# 主流程
# ──────────────────────────────────────────────

def process_one(norm_file: str, usd_file: str, country: str, year: int):
    """处理一组文件：USD表算比值参数，归一化表算绝对量参数。"""
    norm_path = os.path.join(BASE_DIR, norm_file)
    usd_path = os.path.join(EXCEL_DIR, usd_file)
    cap_price = CAPITAL_PRICES[(country, year)]

    print(f"\n{'='*60}")
    print(f"{country}{year}")
    print(f"  USD 源: io_excel/{usd_file}")
    print(f"  归一化源: io-final/{norm_file}")
    print(f"{'='*60}")

    # 解析两个表
    t_usd = parse_excel_io_table(usd_path, country, year)
    t_norm = parse_excel_io_table(norm_path, country, year)

    sectors = t_usd.sectors
    n = len(sectors)
    print(f"  部门: {sectors} (原始: {t_usd.raw_sectors})")

    # 输出目录
    out_dir = os.path.join(BASE_DIR, f"{country}{year}")
    os.makedirs(out_dir, exist_ok=True)

    # ══════════════════════════════════════════════
    # A 类: 从 USD 数据计算比值（保持经济学含义）
    # ══════════════════════════════════════════════
    alpha = compute_alpha_ij(t_usd)
    theta = compute_theta_ij(t_usd)
    beta = compute_beta_j(t_usd)
    theta_c = compute_theta_cj(t_usd)

    save_matrix_csv(os.path.join(out_dir, "alpha_ij.csv"), alpha, sectors)
    save_matrix_csv(os.path.join(out_dir, "theta_ij.csv"), theta, sectors)
    save_vector_csv(os.path.join(out_dir, "beta_j.csv"), beta, sectors,
                    "product_sector_j", "beta_j")
    save_vector_csv(os.path.join(out_dir, "theta_cj.csv"), theta_c, sectors,
                    "product_sector_j", "theta_cj")

    print(f"\n  [A类·USD] alpha_ij 行和: {np.round(alpha.sum(axis=1), 4)}")
    print(f"  [A类·USD] beta_j: {np.round(beta, 4)}  (sum={beta.sum():.6f})")
    print(f"  [A类·USD] theta_ij 对角线: {np.round(np.diag(theta), 4)}")
    print(f"  [A类·USD] theta_cj: {np.round(theta_c, 4)}")

    # ══════════════════════════════════════════════
    # B 类: 绝对量从归一化数据
    # ══════════════════════════════════════════════
    fp = compute_factor_params(t_norm, cap_price)
    export_val = compute_export_value_j(t_norm)
    P_j = np.ones(n)
    P_j_O = np.ones(n)
    E_j = export_val / P_j

    with open(os.path.join(out_dir, "factor_params.csv"), "w") as f:
        f.write("parameter,value\n")
        for k in ["w", "r", "K", "L"]:
            f.write(f"{k},{fp[k]}\n")

    save_vector_csv(os.path.join(out_dir, "P_j.csv"), P_j, sectors,
                    "sector_j", "P_j")
    save_vector_csv(os.path.join(out_dir, "P_j_O.csv"), P_j_O, sectors,
                    "sector_j", "P_j_O")
    save_vector_csv(os.path.join(out_dir, "E_j.csv"), E_j, sectors,
                    "sector_j", "E_j")
    save_vector_csv(os.path.join(out_dir, "ExportValue_j.csv"), export_val, sectors,
                    "sector_j", "ExportValue_j")

    print(f"\n  [B类·归一化] w={fp['w']}, r={fp['r']:.6f}, K={fp['K']:.4f}, L={fp['L']:.6f}")
    print(f"  [B类·归一化] ExportValue_j: {np.round(export_val, 6)}")

    # ══════════════════════════════════════════════
    # C 类: gamma 从 USD 比值导出，其余外生
    # ══════════════════════════════════════════════
    gamma_ij = compute_gamma_ij(theta)
    gamma_cj = compute_gamma_cj(theta_c)
    A_i = np.ones(n)
    rho_cj = np.full(n, RHO_CJ_DEFAULT)
    rho_ij = np.full((n, n), RHO_IJ_DEFAULT)
    import_cost = np.full(n, IMPORT_COST_DEFAULT)

    save_matrix_csv(os.path.join(out_dir, "gamma_ij.csv"), gamma_ij, sectors)
    save_vector_csv(os.path.join(out_dir, "gamma_cj.csv"), gamma_cj, sectors,
                    "sector_j", "gamma_cj")
    save_vector_csv(os.path.join(out_dir, "A_i.csv"), A_i, sectors,
                    "sector_i", "A_i")
    save_vector_csv(os.path.join(out_dir, "rho_cj.csv"), rho_cj, sectors,
                    "sector_j", "rho_cj")
    save_matrix_csv(os.path.join(out_dir, "rho_ij.csv"), rho_ij, sectors)
    save_vector_csv(os.path.join(out_dir, "import_cost.csv"), import_cost, sectors,
                    "sector_j", "import_cost")
    save_time_series_csv(os.path.join(out_dir, "Export_t.csv"), N_PERIODS, sectors)
    save_time_series_csv(os.path.join(out_dir, "P_t_O.csv"), N_PERIODS, sectors)
    save_vector_csv(os.path.join(out_dir, "Export_base_j.csv"), export_val, sectors,
                    "sector_j", "Export_base_j")
    save_vector_csv(os.path.join(out_dir, "P_t_O_base_j.csv"), P_j_O, sectors,
                    "sector_j", "P_t_O_base_j")

    print(f"\n  [C类·USD] gamma_ij 对角线: {np.round(np.diag(gamma_ij), 4)}")
    print(f"  [C类·USD] gamma_cj: {np.round(gamma_cj, 4)}")

    # ── metadata ──
    metadata = {
        "country": country,
        "year": year,
        "source_files": {
            "ratio_params_from": f"io_excel/{usd_file}",
            "quantity_params_from": f"io-final/{norm_file}",
        },
        "sectors": sectors,
        "raw_sector_codes": t_usd.raw_sectors,
        "capital_factor_price": cap_price,
        "n_sectors": n,
        "methodology": {
            "ratio_params": "alpha_ij, theta_ij, beta_j, theta_cj, gamma_ij, gamma_cj — computed from pre-normalization USD data to preserve economic ratios",
            "quantity_params": "factor_params, ExportValue_j, E_j — computed from min-max normalized data",
            "exogenous": "P_j=1, P_j_O=1, A_i=1, rho_cj=0.95",
        },
    }
    with open(os.path.join(out_dir, "metadata.json"), "w", encoding="utf-8") as f:
        json.dump(metadata, f, ensure_ascii=False, indent=2)

    print(f"\n  已保存 {len(os.listdir(out_dir))} 个文件到 io-final/{country}{year}/")
    return sectors, alpha, theta, beta, theta_c, fp, export_val


def main():
    print("=" * 60)
    print("IO 表 → 模型参数计算")
    print("  比值参数 ← USD 原始数据 (保持经济学含义)")
    print("  绝对量参数 ← 归一化数据")
    print("=" * 60)

    results = {}
    for norm_file, usd_file, country, year in FILES:
        results[(country, year)] = process_one(norm_file, usd_file, country, year)

    # ── 汇总对比 ──
    print(f"\n\n{'='*80}")
    print("参数汇总对比")
    print("=" * 80)

    print(f"\nalpha_ij 行和 (应 < 1):")
    print(f"{'':16} {'A01':>8} {'RUP':>8} {'RMID':>8} {'ELEC':>8} {'TEQ':>8}")
    for (c, y), (s, alpha, *_) in results.items():
        rs = alpha.sum(axis=1)
        print(f"  {c}{y}:       " + "  ".join(f"{v:>8.4f}" for v in rs))

    print(f"\nbeta_j (sum=1):")
    print(f"{'':16} {'A01':>8} {'RUP':>8} {'RMID':>8} {'ELEC':>8} {'TEQ':>8} {'sum':>8}")
    for (c, y), (s, _, _, beta, *_) in results.items():
        print(f"  {c}{y}:       " + "  ".join(f"{v:>8.4f}" for v in beta) + f"  {beta.sum():>8.4f}")

    print(f"\ntheta_cj (消费端国产份额):")
    print(f"{'':16} {'A01':>8} {'RUP':>8} {'RMID':>8} {'ELEC':>8} {'TEQ':>8}")
    for (c, y), (s, _, _, _, theta_c, *_) in results.items():
        print(f"  {c}{y}:       " + "  ".join(f"{v:>8.4f}" for v in theta_c))

    print(f"\n要素参数:")
    print(f"{'':16} {'w':>8} {'r':>10} {'K':>12} {'L':>12}")
    for (c, y), (_, _, _, _, _, fp, _) in results.items():
        print(f"  {c}{y}:       {fp['w']:>8.2f} {fp['r']:>10.6f} {fp['K']:>12.4f} {fp['L']:>12.6f}")

    print(f"\n{'='*80}")
    print("全部完成!")
    for _, _, country, year in FILES:
        out_dir = os.path.join(BASE_DIR, f"{country}{year}")
        n_files = len(os.listdir(out_dir))
        print(f"  io-final/{country}{year}/  ({n_files} files)")


if __name__ == "__main__":
    main()
